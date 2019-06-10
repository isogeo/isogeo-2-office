# -*- coding: UTF-8 -*-
#! python3

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# config headers
cols_v = {
    "Titre": (35,),  # A
    "Nom": (35,),  # B
    "Résumé": (35,),  # C
    "Emplacement": (35,),  # D
    "Groupe de travail": (35,),  # E
    "Mots-clés": (35,),  # F
    "Thématique(s) INSPIRE": (35,),  # G
    "Conformité INSPIRE": (35,),  # H
    "Contexte de collecte": (35,),  # I
    "Méthode de collecte": (35,),  # J
    "Début de validité": (35,),  # K
    "Fin de validité": (35,),  # L
    "Fréquence de mise à jour": (35,),  # M
    "Commentaire": (35,),  # N
    "Création": (35,),  # O
    "# mises à jour": (35,),  # P
    "Dernière mise à jour": (35,),  # Q
    "Publication": (35,),  # R
    "Format (version - encodage)": (35,),  # S
    "SRS (EPSG)": (35,),  # T
    "Emprise": (35,),  # U
    "Géométrie": (35,),  # V
    "Résolution": (35,),  # W
    "Echelle": (35,),  # X
    "# Objets": (35,),  # Y
    "# Attributs": (35,),  # Z
    "Attributs (A-Z)": (35,),  # AA
    "Spécifications": (35,),  # AB
    "Cohérence topologique": (35,),  # AC
    "Conditions": (35,),  # AD
    "Limitations": (35,),  # AE
    "# Contacts": (35,),  # AF
    "Points de contact": (35,),  # AG
    "Autres contacts": (35,),  # AH
    "Téléchargeable": (35,),  # AI
    "Visualisable": (35,),  # AJ
    "Autres": (35,),  # AK
    "Editer": (35,),  # AL
    "Consulter": (35,),  # AM
    "MD - ID": (35,),  # AN
    "MD - Création": (35,),  # AO
    "MD - Modification": (35,),  # AP
    "MD - Langue": (35,),  # AQ
}

# workookk and worksheet instances
wb = Workbook()
ws = wb.create_sheet("i2o_worksheet_struct")

x = 1
for i in cols_v:
    _ = ws.cell(column=x, row=1, value=i)
    ws.column_dimensions[get_column_letter(x)].width = 75
    x += 1


wb.save("test_xl_worksheet_struct.xlsx")
