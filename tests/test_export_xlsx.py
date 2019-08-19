# -*- coding: UTF-8 -*-
#! python3

"""
    Usage from the repo root folder:

    ```python
    # for whole test
    python -m unittest tests.test_export_xlsx
    # for specific
    python -m unittest tests.test_export_xlsx.TestExportXLSX.test_metadata_export
    ```
"""

# #############################################################################
# ########## Libraries #############
# ##################################

# Standard library
import json
from os import environ, path
from tempfile import mkstemp
import unittest

# 3rd party
from isogeo_pysdk import Metadata
from openpyxl import Workbook

# target
from modules import Isogeo2xlsx
from modules import IsogeoFormatter
from modules import IsogeoStats

# #############################################################################
# ######## Globals #################
# ##################################

# API access
app_id = environ.get("ISOGEO_API_DEV_ID")
app_secret = environ.get("ISOGEO_API_DEV_SECRET")

# #############################################################################
# ########## Classes ###############
# ##################################


class TestExportXLSX(unittest.TestCase):
    """Test export to Microsoft Excel XLSX."""

    def setUp(self):
        """Executed before each test."""
        # API response samples
        self.tags_sample_types_all = path.normpath(
            r"tests/fixtures/api_response_tests_tags_types_all.json"
        )
        self.tags_sample_types_noRaster = path.normpath(
            r"tests/fixtures/api_response_tests_tags_types_noRaster.json"
        )
        self.tags_sample_types_noVector = path.normpath(
            r"tests/fixtures/api_response_tests_tags_types_noVector.json"
        )
        self.tags_sample_types_noService = path.normpath(
            r"tests/fixtures/api_response_tests_tags_types_noService.json"
        )
        self.tags_sample_types_noResource = path.normpath(
            r"tests/fixtures/api_response_tests_tags_types_noResource.json"
        )
        self.search_all_includes = path.normpath(
            r"tests/fixtures/api_search_complete.json"
        )

        # target class instanciation
        self.out_wb = Isogeo2xlsx(lang="FR", url_base="https://open.isogeo.com/s")

    def tearDown(self):
        """Executed after each test."""
        pass

    # -- Workbook structure ---------------------------------------------------
    def test_workbook_subclass(self):
        """Test module subclass."""
        self.assertIsInstance(self.out_wb, Workbook)
        self.assertEqual(len(self.out_wb.worksheets), 0)

    def test_output_attributes(self):
        """Test output workbook attributes."""
        # attributes
        self.assertTrue(hasattr(self.out_wb, "cols_v"))
        self.assertTrue(hasattr(self.out_wb, "cols_r"))
        self.assertTrue(hasattr(self.out_wb, "cols_s"))
        self.assertTrue(hasattr(self.out_wb, "cols_rz"))
        self.assertTrue(hasattr(self.out_wb, "cols_fa"))
        self.assertTrue(hasattr(self.out_wb, "url_base"))
        self.assertTrue(hasattr(self.out_wb, "dates_fmt"))
        self.assertTrue(hasattr(self.out_wb, "locale_fmt"))
        self.assertTrue(hasattr(self.out_wb, "stats"))
        self.assertTrue(hasattr(self.out_wb, "tr"))
        self.assertTrue(hasattr(self.out_wb, "fmt"))

        # values
        self.assertEqual(self.out_wb.url_base, "https://open.isogeo.com/s")
        self.assertIsInstance(self.out_wb.fmt, IsogeoFormatter)
        self.assertIsInstance(self.out_wb.stats, IsogeoStats)

        # languages variations
        out_wb_fr = Isogeo2xlsx(lang="FR", url_base="https://open.isogeo.com/s")
        self.assertEqual(out_wb_fr.dates_fmt, "DD/MM/YYYY")
        self.assertEqual(out_wb_fr.locale_fmt, "fr_FR")

        out_wb_en = Isogeo2xlsx(lang="EN", url_base="https://open.isogeo.com/s")
        self.assertEqual(out_wb_en.dates_fmt, "YYYY/MM/DD")
        self.assertEqual(out_wb_en.locale_fmt, "uk_UK")

    def test_output_structure_all_types_basic(self):
        """Test output workbook worksheets with all types tags."""
        # load tags fixtures
        with open(self.tags_sample_types_all, "r") as f:
            search = json.loads(f.read())
        # run
        self.out_wb.set_worksheets(auto=search.get("tags").keys())
        self.assertEqual(len(self.out_wb.worksheets), 4)
        self.assertIn("Raster", self.out_wb.sheetnames)
        self.assertIn("Services", self.out_wb.sheetnames)
        self.assertIn("Ressources", self.out_wb.sheetnames)
        self.assertIn("Vecteurs", self.out_wb.sheetnames)

    def test_output_structure_all_types_attributes(self):
        """Test output workbook worksheets with all types tags and attributes."""
        # load tags fixtures
        with open(self.tags_sample_types_all, "r") as f:
            search = json.loads(f.read())
        # run
        self.out_wb.set_worksheets(auto=search.get("tags").keys(), attributes=1)
        self.assertEqual(len(self.out_wb.worksheets), 5)
        self.assertIn("Raster", self.out_wb.sheetnames)
        self.assertIn("Services", self.out_wb.sheetnames)
        self.assertIn("Ressources", self.out_wb.sheetnames)
        self.assertIn("Vecteurs", self.out_wb.sheetnames)
        self.assertIn("Attributs", self.out_wb.sheetnames)

    def test_output_structure_all_types_options(self):
        """Test output workbook worksheets with all types tags and other options."""
        # load tags fixtures
        with open(self.tags_sample_types_all, "r") as f:
            search = json.loads(f.read())
        # run
        self.out_wb.set_worksheets(
            auto=search.get("tags").keys(),
            attributes=1,
            inspire=1,
            dashboard=1,
            fillfull=1,
        )
        self.assertEqual(len(self.out_wb.worksheets), 8)
        self.assertIn("Raster", self.out_wb.sheetnames)
        self.assertIn("Services", self.out_wb.sheetnames)
        self.assertIn("Ressources", self.out_wb.sheetnames)
        self.assertIn("Vecteurs", self.out_wb.sheetnames)
        self.assertIn("Attributs", self.out_wb.sheetnames)
        self.assertIn("Tableau de bord", self.out_wb.sheetnames)
        self.assertIn("Progression catalogage", self.out_wb.sheetnames)

    def test_output_structure_noVector_attributes_bad(self):
        """Test output workbook worksheets with all types tags."""
        # load tags fixtures
        with open(self.tags_sample_types_noVector, "r") as f:
            search = json.loads(f.read())
        # run
        self.out_wb.set_worksheets(auto=search.get("tags").keys(), attributes=1)
        self.assertEqual(len(self.out_wb.worksheets), 3)
        self.assertIn("Raster", self.out_wb.sheetnames)
        self.assertIn("Services", self.out_wb.sheetnames)
        self.assertIn("Ressources", self.out_wb.sheetnames)
        self.assertNotIn("Vecteurs", self.out_wb.sheetnames)
        self.assertNotIn("Attributs", self.out_wb.sheetnames)

    def test_output_structure_noRaster_attributes(self):
        """Test output workbook worksheets with all types tags."""
        # load tags fixtures
        with open(self.tags_sample_types_noRaster, "r") as f:
            search = json.loads(f.read())
        # run
        self.out_wb.set_worksheets(auto=search.get("tags").keys(), attributes=1)
        self.assertEqual(len(self.out_wb.worksheets), 4)
        self.assertNotIn("Raster", self.out_wb.sheetnames)
        self.assertIn("Services", self.out_wb.sheetnames)
        self.assertIn("Ressources", self.out_wb.sheetnames)
        self.assertIn("Vecteurs", self.out_wb.sheetnames)
        self.assertIn("Attributs", self.out_wb.sheetnames)

    def test_output_structure_noService(self):
        """Test output workbook worksheets with all types tags."""
        # load tags fixtures
        with open(self.tags_sample_types_noService, "r") as f:
            search = json.loads(f.read())
        # run
        self.out_wb.set_worksheets(auto=search.get("tags").keys())
        self.assertEqual(len(self.out_wb.worksheets), 3)
        self.assertIn("Raster", self.out_wb.sheetnames)
        self.assertNotIn("Services", self.out_wb.sheetnames)
        self.assertIn("Ressources", self.out_wb.sheetnames)
        self.assertIn("Vecteurs", self.out_wb.sheetnames)

    def test_output_structure_noResource(self):
        """Test output workbook worksheets with all types tags."""
        # load tags fixtures
        with open(self.tags_sample_types_noResource, "r") as f:
            search = json.loads(f.read())
        # run
        self.out_wb.set_worksheets(auto=search.get("tags").keys())
        self.assertEqual(len(self.out_wb.worksheets), 3)
        self.assertIn("Raster", self.out_wb.sheetnames)
        self.assertIn("Services", self.out_wb.sheetnames)
        self.assertNotIn("Ressources", self.out_wb.sheetnames)
        self.assertIn("Vecteurs", self.out_wb.sheetnames)

    def test_output_structure_bad(self):
        """Test output workbook worksheets with all types tags."""
        # run
        with self.assertRaises(TypeError):
            self.out_wb.set_worksheets(auto=["type:resource", "type:service"])

    # -- Export ---------------------------------------------------------------
    def test_metadata_export(self):
        """Test search results export"""
        # temp output file
        out_xlsx = mkstemp(prefix="i2o_test_xlsx_")
        # load tags fixtures
        with open(self.search_all_includes, "r") as f:
            search = json.loads(f.read())
        # add worksheets
        self.out_wb.set_worksheets(auto=search.get("tags").keys())
        # run
        for md in search.get("results"):
            # clean invalid attributes
            md["coordinateSystem"] = md.pop("coordinate-system", list)
            md["featureAttributes"] = md.pop("feature-attributes", list)
            # load metadata
            metadata = Metadata(**md)
            self.out_wb.store_metadatas(metadata)
        # save
        self.out_wb.save(out_xlsx[1] + ".xlsx")

    def test_metadata_bad_no_dict(self):
        """Test metadata not as a dict."""
        # run
        with self.assertRaises(TypeError):
            self.out_wb.store_metadatas(
                ["_id", "azertyqwerty", "title", "fixturing me"]
            )

    # def test_metadata_bad_type(self):
    #     """Test bad metadata type."""
    #     # run
    #     with self.assertRaises(TypeError):
    #         self.out_wb.store_metadatas({"_id": "azertyqwerty",
    #                                      "type": "datasetEtMatch"})


# #############################################################################
# ##### Stand alone program ########
# ##################################
if __name__ == "__main__":
    unittest.main()
