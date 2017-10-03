# -*- coding: UTF-8 -*-
#!/usr/bin/env python
from __future__ import (absolute_import, print_function, unicode_literals)
# ----------------------------------------------------------------------------
# Purpose:      Get metadatas from an Isogeo share and store it into files
#
# Author:       Julien Moura (@geojulien)
#
# Python:       2.7.x
# Created:      18/12/2015
# Updated:      22/01/2016
# ---------------------------------------------------------------------------

# ############################################################################
# ########## Libraries #############
# ##################################

# Standard library
from datetime import datetime
import gettext  # localization
import logging
from logging.handlers import RotatingFileHandler
from os import path, walk
import platform  # about operating systems
from sys import argv, exit
from tempfile import mkdtemp
from tkinter.filedialog import askdirectory, askopenfilename
from tkinter.messagebox import showerror as avert
from tkinter import Tk, Image, PhotoImage
from tkinter import IntVar, StringVar, ACTIVE, DISABLED, VERTICAL
from tkinter.ttk import Label, Button, Checkbutton
from tkinter.ttk import Labelframe, Progressbar, Separator, Style
from zipfile import ZipFile

# 3rd party library
from docxtpl import DocxTemplate
from isogeo_pysdk import Isogeo, __version__ as pysdk_version
import openpyxl
import requests

# Custom modules
from modules import Isogeo2xlsx
from modules import Isogeo2docx
from modules import IsogeoAppAuth
from modules import IsogeoStats
from modules import CheckNorris
from modules import isogeo2office_utils

# UI submodules
from modules import FrameExcel
from modules import FrameGlobal
from modules import FrameWord
from modules import FrameXml
from modules import ToolTip

# ############################################################################
# ########## Global ################
# ##################################

# VERSION
_version = "1.5.7"

# LOG FILE ##
logger = logging.getLogger("isogeo2office")
logging.captureWarnings(True)
logger.setLevel(logging.DEBUG)
log_form = logging.Formatter("%(asctime)s || %(levelname)s "
                             "|| %(module)s || %(lineno)s || %(message)s")
logfile = RotatingFileHandler("LOG_isogeo2office.log", "a", 5000000, 1)
logfile.setLevel(logging.DEBUG)
logfile.setFormatter(log_form)
logger.addHandler(logfile)
logger.info('\n')
logger.info('================ Isogeo to office ===============')

# ############################################################################
# ########## Classes ###############
# ##################################


class Isogeo2office(Tk):
    """Main Class for Isogeo to Office."""

    # attributes and global actions
    logger.info('OS: {0}'.format(platform.platform()))
    logger.info('Version: {0}'.format(_version))
    logger.info('Isogeo PySDK version: {0}'.format(pysdk_version))

    def __init__(self, ui_launcher=1, settings_file=r"settings.ini"):
        """Initiliazing isogeo2office with or without UI."""
        # Invoke Check Norris & utils
        checker = CheckNorris()
        self.utils = isogeo2office_utils()
        self.stats = IsogeoStats()

        # ------------ Settings ----------------------------------------------
        self.settings = self.utils.settings_load(path.abspath(settings_file))
        self.app_id = self.settings.get("auth").get("app_id")
        self.app_secret = self.settings.get("auth").get("app_secret")
        self.client_lang = self.settings.get("global").get("def_codelang", "FR")

        # ------------ Localization ------------------------------------------
        if self.client_lang == "FR":
            lang = gettext.translation("isogeo2office", localedir="i18n",
                                       languages=["fr_FR"], codeset="Latin1")
            lang.install()
        else:
            lang = gettext
            lang.install("isogeo2office", localedir="i18n")
        logger.info(u"Language applied: {}".format(_(u"English")))

        # ------------ UI or not to UI ---------------------------------------
        if not ui_launcher:
            self.no_ui_launcher()
        else:
            pass

        # ------------ Internet Connection -----------------------------------
        if not checker.check_internet_connection():
            logger.error('Internet connection required: check your settings.')
            avert(_("Internet connection failed"),
                  _("An Internet connection is required to use Isogeo API"
                    "\n(https://v1.api.isogeo.com:443)."))
            exit()
        else:
            pass

        # ------------ Isogeo authentication -------------------------------
        try:
            self.isogeo = Isogeo(client_id=self.app_id,
                                 client_secret=self.app_secret,
                                 lang=self.client_lang)
            self.token = self.isogeo.connect()
        except:
            # if id/secret doesn't work, ask for a new one
            prompter = IsogeoAppAuth(prev_id=self.app_id,
                                     prev_secret=self.app_secret,
                                     lang=lang)
            prompter.mainloop()
            # check response
            if len(prompter.li_dest) < 2:
                logger.error(u"API authentication form returned nothing.")
                exit()
            else:
                pass

            self.app_id = prompter.li_dest[0]
            self.app_secret = prompter.li_dest[1]
            self.isogeo = Isogeo(client_id=self.app_id,
                                 client_secret=self.app_secret,
                                 lang=self.client_lang)
            self.token = self.isogeo.connect()

        # debug logs
        logger.debug("API: " + self.isogeo.get_isogeo_version(component="api"))
        logger.debug("APP: " + self.isogeo.get_isogeo_version(component="app"))
        logger.debug("DB: " + self.isogeo.get_isogeo_version(component="db"))
        # ------------ Isogeo search & shares --------------------------------
        self.search_results = self.isogeo.search(self.token,
                                                 page_size=0,
                                                 whole_share=0)
        self.shares = self.isogeo.shares(self.token)
        self.shares_info = self.get_shares_info()

        # ------------ Variables ---------------------------------------------

        # ------------ UI ----------------------------------------------------
        Tk.__init__(self)
        self.title("isogeo2office - {0}".format(_version))
        icon = Image("photo", file=r"img/favicon_isogeo.gif")
        self.call("wm", "iconphoto", self._w, icon)
        self.style = Style(self).theme_use("vista")
        self.resizable(width=False, height=False)
        self.focus_force()
        self.msg_bar = StringVar(self)
        self.msg_bar.set(_(u"Pick your options and push the launch button"))

        # styling
        btn_style_err = Style(self)
        btn_style_err.configure('Error.TButton', foreground='Red')

        # cbb_style_err = Style(self)
        # cbb_style_err.configure('TCombobox', foreground='Red')

        # fields validation
        fields_validators = {
            "val_uid": (self.register(self.utils.entry_validate_uid),
                        '%d', '%i', '%P', '%s', '%S', '%v', '%V', '%W'),
            "val_date": (self.register(self.utils.entry_validate_date),
                         '%d', '%i', '%P', '%s', '%S', '%v', '%V', '%W')
        }

        # Frames and main widgets
        fr_actions = FrameGlobal(self, main_path="", lang=lang)
        fr_isogeo = Labelframe(self, name="isogeo", text="Isogeo")
        self.fr_excel = FrameExcel(self, main_path="", lang=lang)
        self.fr_word = FrameWord(self, main_path="", lang=lang,
                                 validators=fields_validators)
        self.fr_xml = FrameXml(self, main_path="", lang=lang,
                               validators=fields_validators)
        fr_process = Labelframe(self, name="process", text="Launch")
        self.status_bar = Label(self, textvariable=self.msg_bar, anchor='w',
                                foreground='DodgerBlue')
        self.progbar = Progressbar(self,
                                   orient="horizontal")

        fr_actions.grid(row=0, column=1, columnspan=2, padx=2, pady=4, sticky="WE")
        fr_isogeo.grid(row=1, column=1, columnspan=5, padx=2, pady=4, sticky="WE")
        self.fr_excel.grid(row=2, column=1, padx=2, pady=4, sticky="WE")
        self.fr_word.grid(row=3, column=1, padx=2, pady=4, sticky="WE")
        self.fr_xml.grid(row=4, column=1, padx=2, pady=4, sticky="WE")
        fr_process.grid(row=5, column=1, padx=2, pady=4, sticky="WE")
        self.status_bar.grid(row=6, column=1, padx=2, pady=2, sticky="WE")
        self.progbar.grid(row=7, column=1, sticky="WE")

        # --------------------------------------------------------------------

        # ## GLOBAL ##
        self.app_metrics = StringVar(fr_isogeo)
        self.oc_msg = StringVar(fr_isogeo)
        self.url_input = StringVar(fr_isogeo)

        # logo
        self.logo_isogeo = PhotoImage(master=fr_isogeo,
                                      file=r'img/logo_isogeo.gif')
        logo_isogeo = Label(fr_isogeo, borderwidth=2, image=self.logo_isogeo)
        ToolTip(logo_isogeo, message="Logo Isogeo")

        # metrics
        self.app_metrics.set(_("{}\n{} metadata in {} share(s)"
                               " owned by {} workgroup(s).")
                             .format(self.shares_info[-1].get("name"),
                                     self.search_results.get('total'),
                                     len(self.shares),
                                     len(self.shares_info[1])))
        lb_app_metrics = Label(fr_isogeo,
                               textvariable=self.app_metrics)

        # OpenCatalog check
        self.lb_input_oc = Label(fr_isogeo,
                                 textvariable=self.oc_msg)

        if len(self.shares_info[2]) != 0:
            logger.error("Any OpenCatalog found among the shares")
            self.oc_msg.set(_("{} shares don't have any OpenCatalog."
                              "\nAdd OpenCatalog to every share,"
                              "\nthen reboot isogeo2office.").format(len(self.shares_info[2])))
            self.msg_bar.set(_("Error: some shares don't have OpenCatalog"
                               "activated. Fix it first."))
            self.status_bar.config(foreground='Red')
            fr_actions.btn_open_shares.configure(command=lambda: self.utils.open_urls(self.shares_info[2]))
            status_launch = DISABLED
        elif len(self.shares) != len(self.shares_info[1]):
            logger.error("More than one share by workgroup")
            self.oc_msg.set(_("Too much shares by workgroup."
                              "\nPlease red button to fix it"
                              "\nthen reboot isogeo2office.")
                            .format(len(self.shares) - len(self.shares_info[1])))
            self.msg_bar.set(_("Error: more than one share by worgroup."
                               " Click on Admin button to fix it."))
            self.status_bar.config(foreground='Red')
            fr_actions.btn_open_shares.configure(text=_("Fix the shares"),
                                                 command=lambda: self.utils.open_urls(self.shares_info[3]),
                                                 style="Error.TButton")
            status_launch = DISABLED
        else:
            logger.info("All shares have an OpenCatalog")
            self.oc_msg.set(_("Configuration OK."))
            li_oc = [share[3] for share in self.shares_info[0]]
            fr_actions.btn_open_shares.configure(command=lambda: self.utils.open_urls(li_oc))
            status_launch = ACTIVE

        # settings
        # for unicode symbols: https://www.w3schools.com/charsets/ref_utf_symbols.asp
        # griding widgets
        logo_isogeo.grid(row=1, rowspan=3,
                         column=0, padx=2,
                         pady=2, sticky="W")
        Separator(fr_isogeo, orient=VERTICAL).grid(row=1, rowspan=3,
                                                   column=1, padx=2,
                                                   pady=2, sticky="NSE")
        lb_app_metrics.grid(row=1, column=2, rowspan=3, sticky="NWE")
        self.lb_input_oc.grid(row=2, column=2, sticky="WE")

        # --------------------------------------------------------------------

        # ## EXCEL ##
        # variables
        self.fr_excel.output_name.set(self.settings.get("excel")
                                                   .get("output_name",
                                                        "isogeo2xlsx"))
        self.fr_excel.opt_dashboard.set(self.settings.get("excel")
                                                     .get("opt_dashboard", 1))
        self.fr_excel.opt_attributes.set(self.settings.get("excel")
                                                      .get("opt_attributes", 0))
        self.fr_excel.opt_fillfull.set(self.settings.get("excel")
                                                    .get("opt_fillfull", 0))
        self.fr_excel.opt_inspire.set(self.settings.get("excel")
                                                   .get("opt_inspire", 0))

        # --------------------------------------------------------------------

        # ## WORD ##
        # variables
        self.fr_word.out_prefix.set(self.settings.get("word")
                                                 .get("out_prefix",
                                                      "isogeo2docx"))
        self.fr_word.opt_id.set(self.settings.get("word")
                                             .get("opt_id", 5))
        self.fr_word.opt_date.set(self.settings.get("word")
                                               .get("opt_date", 1))

        # --------------------------------------------------------------------

        # ## XML ##
        # variables
        self.fr_xml.out_prefix.set(self.settings.get("xml")
                                                .get("out_prefix",
                                                     "isogeo2xml"))
        self.fr_xml.opt_id.set(self.settings.get("xml")
                                            .get("opt_id", 5))
        self.fr_xml.opt_date.set(self.settings.get("xml")
                                              .get("opt_date", 1))

        self.fr_xml.opt_zip.set(self.settings.get("xml")
                                             .get("opt_zip", 1))

        # --------------------------------------------------------------------

        # ## PROCESS ##
        # variables
        self.opt_excel = IntVar(fr_process,
                                int(self.settings.get('excel')
                                    .get('excel_opt', 0))
                                )
        self.opt_word = IntVar(fr_process,
                               int(self.settings.get('word')
                                   .get('word_opt', 0)))
        self.opt_xml = IntVar(fr_process,
                              int(self.settings.get('xml')
                                  .get('xml_opt', 0)))
        # self.out_folder_path = self.settings.get("basics").get("out_folder",
        #                                                        "output")
        try:
            self.out_fold_path = StringVar(fr_process,
                                           path.relpath(self.settings.get('global')
                                                        .get("out_folder")))
        except ValueError as e:
            logger.debug(e)
            self.out_fold_path = StringVar(fr_process,
                                           path.relpath("output"))

        # logo
        self.logo_process = PhotoImage(master=fr_process,
                                       file=r'img/logo_process.gif')
        logo_process = Label(fr_process, borderwidth=2,
                             image=self.logo_process)

        # options
        caz_go_excel = Checkbutton(fr_process,
                                   text=_(u'Export the whole shares into an Excel worksheet'),
                                   variable=self.opt_excel)

        caz_go_word = Checkbutton(fr_process,
                                  text=_(u'Export each metadata into a Word file'),
                                  variable=self.opt_word)

        caz_go_xml = Checkbutton(fr_process,
                                 text=_(u'Export each metadata into a XML ISO-19139'),
                                 variable=self.opt_xml)

        # output folder
        lb_out_fold_title = Label(fr_process, text=_("Output folder: "))
        lb_out_fold_var = Label(fr_process,
                                textvariable=self.out_fold_path)
        btn_out_fold_path_browse = Button(fr_process,
                                          text="{} {}".format("\U000026D7 ", _("Browse")),
                                          command=lambda: self.set_out_folder_path(self.out_fold_path.get()))

        btn_out_fold_open = Button(fr_process,
                                   text=u"\U000021F6 " + _("Open"),
                                   command=lambda: self.utils.open_dir_file(self.out_fold_path.get()))

        # launcher
        self.btn_go = Button(fr_process,
                             text="\U00002708 " + _("Launch"),
                             command=lambda: self.process(),
                             state=status_launch)

        # griding widgets
        logo_process.grid(row=1, rowspan=5,
                          column=0, padx=2,
                          pady=2, sticky="W")
        Separator(fr_process, orient=VERTICAL).grid(row=1, rowspan=5,
                                                    column=1, padx=2,
                                                    pady=2, sticky="NSE")
        caz_go_word.grid(row=2, column=2, columnspan=4, padx=2, pady=2, sticky="W")
        caz_go_excel.grid(row=3, column=2, columnspan=5, padx=2, pady=2, sticky="W")
        caz_go_xml.grid(row=4, column=2, columnspan=5, padx=2, pady=2, sticky="W")
        lb_out_fold_title.grid(row=5, column=2,
                               padx=2, pady=2, sticky="W")
        lb_out_fold_var.grid(row=5, column=3, columnspan=2,
                             padx=2, pady=2, sticky="WE")
        btn_out_fold_path_browse.grid(row=5, column=5,
                                      padx=2, pady=2, sticky="E")
        btn_out_fold_open.grid(row=5, column=6,
                               padx=2, pady=2, sticky="E")
        self.btn_go.grid(row=6, column=2, columnspan=5,
                         padx=2, pady=2, sticky="WE")

        logger.info("Main UI instanciated & displayed")

# ----------------------------------------------------------------------------

    def set_out_folder_path(self, out_folder_path="output"):
        """Open a popup to select a folder and store it."""
        self.btn_go.config(state=DISABLED)  # disable launch in the meanwhile
        foldername = askdirectory(parent=self,
                                  initialdir=path.realpath(out_folder_path),
                                  mustexist=True,
                                  title=_("Select output folder"))
        # check if a folder has been choosen
        if foldername:
            self.out_fold_path.set(path.relpath(foldername))
        else:
            avert(title=_("No folder selected"),
                  message=_("You must select an output folder."))

        self.btn_go.config(state=ACTIVE)
        # end of function
        return foldername

    def get_input_xl(self):
        """Get the path of the input Excel file with a browse dialog."""
        self.input_xl = askopenfilename(parent=self,
                                        filetypes=[("Excel 2010", "*.xlsx")],
                                        title=_(u"Pick the Excel to merge"))

        # testing file choosen
        if self.input_xl:
            print(self.input_xl)
            pass
        elif path.splittext(self.input_xl)[1] != ".xlsx":
            print("Invalid format")
        else:
            print(_(u'Any file selected'))
            return

        # get headers names
        xlsx_in = openpyxl.load_workbook(filename=self.input_xl,
                                         read_only=True,
                                         guess_types=True,
                                         use_iterators=True)
        ws1 = xlsx_in.worksheets[0]  # ws = première feuille
        cols_names = [ws1.cell(row=ws1.min_row, column=col).value
                      for col in range(1, ws1.max_column)]

        # end of method
        return

    def ui_switch_xljoiner(self):
        """Enable/disable the form for input xl to join."""
        if self.opt_xl_join.get():
            self.fr_input_xl_join.pack()
        else:
            self.fr_input_xl_join.pack_forget()
        # end of function
        return

# ----------------------------------------------------------------------------

    def ui_settings_prompt(self, lang):
        """Get Isogeo settings from another form."""
        prompter = IsogeoAppAuth(prev_id=self.app_id,
                                 prev_secret=self.app_secret,
                                 lang=lang)
        prompter.mainloop()
        # check response
        if len(prompter.li_dest) < 2:
            logger.error(u"API authentication form returned nothing.")
            exit()
            return 0
        elif len(prompter.li_dest) == 2 \
             and self.app_id == prompter.li_dest[0]\
             and self.app_secret == prompter.li_dest[1]:
             logger.info(u"Auth Id and Secret have not changed.")
             return 0
        else:
            pass

        # set new auth settings
        self.app_id = prompter.li_dest[0]
        self.app_secret = prompter.li_dest[1]

        # end of method
        return 1

    def get_shares_info(self):
        """Get Isogeo shares informations from application."""
        # variables
        li_oc = []
        li_owners = []
        li_without_oc = []
        li_too_shares = []
        # parsing
        for share in self.shares:
            # Share caracteristics
            share_name = share.get("name")
            creator_name = share.get("_creator").get("contact").get("name")
            creator_id = share.get("_creator").get("_tag")[6:]
            share_url = "https://app.isogeo.com/groups/{}/admin/shares/{}"\
                        .format(creator_id, share.get("_id"))

            # check if there is only a share per workgroup
            if creator_id in li_owners:
                logger.error("This workgroup has more than 1 share to this application:"
                             " https://app.isogeo.com/groups/{}".format(creator_id))
                li_too_shares.append(share_url)
            else:
                # add to shares owners list
                li_owners.append(creator_id)
                pass

            # OpenCatalog URL construction
            share_details = self.isogeo.share(self.token, share_id=share.get("_id"))
            url_oc = "https://open.isogeo.com/s/{}/{}".format(share.get("_id"),
                                                              share_details.get("urlToken"))

            # Testing URL
            request = requests.get(url_oc)
            if request.status_code != 200:
                logger.info("No OpenCatalog set for this share: " + share_url)
                li_without_oc.append(share_url)
                continue
            else:
                pass

            # consolidate list of OpenCatalog available
            li_oc.append((share_name, creator_id, creator_name,
                          share_url, url_oc))
        # get app properties
        first_share = share.get("applications")[0]
        app = {"creation_date": first_share.get("_created"),
               "last_update": first_share.get("_modified"),
               "name": first_share.get("name"),
               "type": first_share.get("type"),
               "url": first_share.get("url")
               }
        logger.info("Isogeo - Shares informations retrieved.")
        # end of method
        return li_oc, set(li_owners), li_without_oc, li_too_shares, app

# ----------------------------------------------------------------------------

    def process(self):
        """Process export according to options set."""
        if not (self.opt_excel.get() + self.opt_word.get() + self.opt_xml.get()):
            self.msg_bar.set(_("Error: at least one export option required"))
            logger.error("Any export option selected.")
            return
        else:
            pass

        # savings in ini file
        self.utils.settings_save(parent=self,
                                 config_file=path.realpath(r"settings.ini"))

        # prepare Isogeo request
        self.msg_bar.set(_("Fetching Isogeo data..."))
        includes = ["conditions",
                    "contacts",
                    "coordinate-system",
                    "events",
                    "feature-attributes",
                    "keywords",
                    "layers",
                    "limitations",
                    "links",
                    "operations",
                    "serviceLayers",
                    "specifications"]

        self.search_results = self.isogeo.search(self.token,
                                                 sub_resources=includes)
        self.progbar["maximum"] = self.search_results.get("total")
        logger.info("Isogeo - metadatas retrieved.")
        # EXCEL
        if self.opt_excel.get():
            logger.info("Excel - START")
            out_xlsx_path = path.realpath(path.join(self.out_fold_path.get(),
                                                    self.fr_excel.output_name.get() + ".xlsx"))
            self.progbar["value"] = 0
            self.process_excelization(output_filepath=out_xlsx_path)
        else:
            pass

        # WORD
        template_path = path.realpath(path.join(r"templates",
                                                self.fr_word.tpl_input.get()))
        if self.opt_word.get() and path.isfile(template_path):
            self.status_bar.config(foreground='DodgerBlue')
            logger.info("WORD - START")
            self.progbar["value"] = 0
            self.process_wordification()
        elif self.opt_word.get() and self.fr_word.tpl_input.get() == "":
            logger.error("Any template selected.")
            self.msg_bar.set(_("Error: Word template not selected"))
            self.status_bar.config(foreground='Red')
            return
        else:
            self.status_bar.config(foreground='DodgerBlue')
            pass

        if self.opt_xml.get():
            logger.info("XML - START")
            self.progbar["value"] = 0
            self.process_xmlisation()
        else:
            pass

        # end of method
        self.msg_bar.set(_("All tasks are done."))
        logger.info("All tasks are done.")
        return

    def process_excelization(self, output_filepath=r"output/TEST_isogeo2xlsx.xlsx", ui=1):
        """Export metadatas shared into an Excel worksheet."""
        # workbook
        url_oc = [share[4] for share in self.shares_info[0]][0]
        wb = Isogeo2xlsx(lang=self.client_lang, url_base=url_oc)
        wb.set_worksheets(auto=self.search_results.get('tags').keys(),
                          dashboard=self.fr_excel.opt_dashboard.get(),
                          attributes=self.fr_excel.opt_attributes.get(),
                          fillfull=self.fr_excel.opt_fillfull.get(),
                          inspire=self.fr_excel.opt_inspire.get())

        # parsing metadata
        for md in self.search_results.get('results'):
            wb.store_metadatas(md)
            if ui:
                # progression
                md_title = md.get("title", "No title").encode("utf8")
                self.msg_bar.set(_("Processing Excel: {}").format(md_title))
                self.progbar["value"] = self.progbar["value"] + 1
            else:
                pass

        # tunning
        wb.tunning_worksheets()

        # special sheets
        if self.fr_excel.opt_fillfull.get():
            wb.ws_f["A1"] = self.stats.fillfull()
            print(self.stats.md_empty_fields)
        else:
            pass
        if self.fr_excel.opt_dashboard.get():
            # metadata types - pie chart
            pie = self.stats.type_pie(wb.ws_d,
                                      self.search_results.get('total'))
            wb.ws_d.add_chart(pie, "A1")
            # tags - bar charts
            bar = self.stats.keywords_bar(wb.ws_d, self.search_results.get("results"))
            wb.ws_d.add_chart(bar, "A10")
        else:
            pass

        # saving the test file
        # dstamp = datetime.now()
        wb.save(output_filepath)

        logger.info("Excel - DONE {}"
                    .format(output_filepath))
        if ui:
            self.msg_bar.set(_("Export Excel done."))
        else:
            pass
        # end of method
        return

    def process_wordification(self, ui=1):
        """Export each metadata shared to a Word document.

        Transformation is based on the template selected.
        """
        # transformer
        to_docx = Isogeo2docx()

        for md in self.search_results.get("results"):
            # get OpenCatalog related to each metadata
            if len(self.shares) == 1:
                url_oc = [share[4] for share in self.shares_info[0]][0]
            elif len(self.shares) > 1:
                # for share in self.shares_info[0]:
                #     print("\nshare owner: ", share[1])
                #     print("md owner: ", md.get("_creator").get("_id"))
                # print("\n", self.shares_info[0])
                url_oc = [share[4] for share in self.shares_info[0]
                          if share[1] == md.get("_creator").get("_id")][0]
                # print(url_oc)
            elif len(self.shares) != len(self.shares_info[1]):
                logger.error("More than one share by workgroup")
                self.msg_bar.set(_("Error: more than one share by worgroup."
                                   " Open APP or push the button to fix it."))
                self.status_bar.config(foreground='Red')
                return
            else:
                self.status_bar.config(foreground='DodgerBlue')
                pass

            # templating
            tpl = DocxTemplate(path.realpath(path.join(r"templates",
                                                       self.fr_word.tpl_input.get())))
            to_docx.md2docx(tpl, md, url_oc)

            # name
            md_name = md.get("name", md.get("title", "NR"))
            if '.' in md_name:
                md_name = md_name.split(".")[1]
            else:
                pass
            md_name = "_{}".format(md_name)

            # uid
            if self.fr_word.opt_id.get():
                uid = "_{}".format(md.get("_id")[:self.fr_word.opt_id.get()])
            else:
                uid = ""

            # date
            dstamp = datetime.now()
            if self.fr_word.opt_date.get() == 1:
                dstamp = "_{}-{}-{}".format(dstamp.year,
                                            dstamp.month,
                                            dstamp.day)
            elif self.fr_word.opt_date.get() == 2:
                dstamp = "_{}-{}-{}-{}{}{}".format(dstamp.year,
                                                   dstamp.month,
                                                   dstamp.day,
                                                   dstamp.hour,
                                                   dstamp.minute,
                                                   dstamp.second,)
            else:
                dstamp = ""

            # final output name
            out_docx_path = path.join(path.realpath(self.out_fold_path.get()),
                                      "{}{}{}{}.docx"
                                      .format(self.fr_word.out_prefix.get(),
                                              uid,
                                              md_name,
                                              dstamp))
            tpl.save(out_docx_path)
            del tpl

            # progression bar
            # self.msg_bar.set(_("Processing Word: {}").format(md_name[1:]))
            self.msg_bar.set(_("Processing Word: {}")
                             .format(md_name[1:].split(" -")[0]))
            self.progbar["value"] = self.progbar["value"] + 1
            self.update()

        self.msg_bar.set(_("Export Word done."))
        logger.info("Word - DONE: {}".format(out_docx_path))
        # end of method
        return

    def process_xmlisation(self):
        """Exports each metadata into XML ISO 19139"""
        # ZIP or not ZIP
        if not self.fr_xml.opt_zip.get():
            # directly into the output directory
            out_dir = path.realpath(self.out_fold_path.get())
        else:
            # into a temporary directory
            out_dir = mkdtemp(prefix="isogeo_", suffix="_xml")
            logger.info("XML - Temporary directory created: {}".format(out_dir))

        # parsing results
        for md in self.search_results.get("results"):
            # name
            md_title = md.get("title", "NR")
            if '.' in md_title:
                md_title = md_title.split(".")[1]
            else:
                pass
            md_title = "_{}".format(md_title)

            # uid
            if self.fr_xml.opt_id.get():
                uid = "_{}".format(md.get("_id")[:self.fr_xml.opt_id.get()])
            else:
                uid = ""

            # date
            dstamp = datetime.now()
            if self.fr_xml.opt_date.get() == 1:
                dstamp = "_{}-{}-{}".format(dstamp.year,
                                            dstamp.month,
                                            dstamp.day)
            elif self.fr_xml.opt_date.get() == 2:
                dstamp = "_{}-{}-{}-{}{}{}".format(dstamp.year,
                                                   dstamp.month,
                                                   dstamp.day,
                                                   dstamp.hour,
                                                   dstamp.minute,
                                                   dstamp.second,)
            else:
                dstamp = ""

            # final output name
            clean_title = self.utils.clean_filename(md_title.split(" -")[0])
            out_xml_path = path.join(out_dir,
                                     "{}{}{}{}.xml"
                                     .format(self.fr_xml.out_prefix.get(),
                                             uid,
                                             clean_title,
                                             dstamp))

            # export
            xml_stream = self.isogeo.xml19139(self.token, md.get("_id"))
            with open(path.realpath(out_xml_path), 'wb') as out_md:
                for block in xml_stream.iter_content(1024):
                    out_md.write(block)

            logger.info("XML - Exported: {} ({})".format(md.get("name", clean_title),
                                                         md.get("_id")))

            # progression bar
            self.msg_bar.set(_("Processing XML: {}")
                             .format(md_title[1:].split(" -")[0]))
            self.progbar["value"] = self.progbar["value"] + 1
            self.update()

        # ZIP or not ZIP
        if not self.fr_xml.opt_zip.get():
            pass
        else:
            out_zip_path = path.join(self.out_fold_path.get(),
                                     "{}{}.zip"
                                     .format(self.fr_xml.out_prefix.get(),
                                             dstamp))
            final_zip = ZipFile(out_zip_path, "w")
            for root, dirs, files in walk(out_dir):
                for f in files:
                    final_zip.write(path.join(root, f), f)
            final_zip.close()
            logger.info("XML - ZIP: {}".format(out_zip_path))

        self.msg_bar.set(_("Export XML done."))
        logger.info("XML - DONE: {}".format(out_xml_path))
        # end of method
        return

# ----------------------------------------------------------------------------

    def no_ui_launcher(self):
        """Execute scripts without UI and using settings.ini."""
        logger.info('Launched from command prompt')
        # utils
        checker = CheckNorris()
        includes = ["conditions",
                    "contacts",
                    "coordinate-system",
                    "events",
                    "feature-attributes",
                    "keywords",
                    "layers",
                    "limitations",
                    "links",
                    "operations",
                    "serviceLayers",
                    "specifications"]

        try:
            self.isogeo = Isogeo(client_id=self.app_id,
                                 client_secret=self.app_secret,
                                 lang=self.client_lang)
            self.token = self.isogeo.connect()
        except Exception as e:
            print(e)
            exit()

        # get settings
        self.settings = default.settings_load()
        self.search_results = self.isogeo.search(self.token,
                                                 sub_resources=includes)

        # Excel export
        if self.settings.get('excel').get('excel_opt'):
            logger.info("Excel - START")
            out_xlsx_path = path.abspath(path.join(self.settings.get('global')
                                                                .get('out_folder'),
                                                   self.settings.get('excel')
                                                                .get('excel_out') + ".xlsx"))
            self.process_excelization(out_xlsx_path, ui=0)

        else:
            pass

        # # Word export
        # if self.settings.get('basics').get('word_opt'):
        #     logger.info("WORD - START")
        #     self.process_wordification()
        # else:
        #     pass

        # # XML export
        # if self.settings.get('basics').get('xml_opt'):
        #     logger.info("XML - START")
        #     self.process_xmlisation()
        # else:
        #     pass
        exit()
        return

# ###############################################################################
# ###### Stand alone program ########
# ###################################

if __name__ == '__main__':
    """standalone execution
    """
    if len(argv) < 2:
        app = Isogeo2office(ui_launcher=1,
                            settings_file=path.abspath(r"settings_cd_57.ini"))
        app.mainloop()
    elif argv[1] == str(1):
        print("launch UI")
        app = Isogeo2office(ui_launcher=1)
        app.mainloop()
    else:
        print("launch without UI")
        app = Isogeo2office(ui_launcher=0)
        print("Finished!")
