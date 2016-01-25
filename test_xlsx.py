# -*- coding: UTF-8 -*-
#!/usr/bin/env python
from __future__ import (absolute_import, print_function, unicode_literals)
# ----------------------------------------------------------------------------
# Name:         Isogeo
# Purpose:      Get metadatas from an Isogeo share and store it into files
#
# Author:       Julien Moura (@geojulien)
#
# Python:       2.7.x
# Created:      18/12/2015
# Updated:      22/01/2016
# ----------------------------------------------------------------------------

from os import path

from openpyxl import load_workbook


# ----------------------------------------------------------------------------

wb_input = r"input\wb_test.xlsx"


# ouverture du fichier des participants en lecture
xlsx_in = load_workbook(filename=wb_input, read_only=True, guess_types=True, data_only=True)

# noms des onglets
print(xlsx_in.get_sheet_names())

ws = xlsx_in.worksheets[0]  # ws = première feuille

print("Nombre de lignes : ", ws.max_row)
print("Nombre de colonnes : ", ws.max_column)

print("\n", dir(ws), "\n\n")

# col1 = ws.columns[0]
# print(dir(col1))

# print(col1[0].value)

cols_names = [ws.cell(row=ws.min_row, column=col).value for col in range(1, ws.max_column)]

print(cols_names)
