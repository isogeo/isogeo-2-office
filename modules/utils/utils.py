# -*- coding: UTF-8 -*-
#! python3  # noqa: E265

"""
    Name:         Isogeo to Office utilitaries
    Author:       Isogeo
    Python:       3.7.x
"""

# ############################################################################
# ########## Libraries #############
# ##################################

# Standard library
import logging
import re
import subprocess
from datetime import datetime, timedelta
from itertools import zip_longest
from os import R_OK, access, environ, path
from pathlib import Path  # TO DO: replace os.path by pathlib
from sys import platform as opersys
from time import sleep
from urllib.request import getproxies
from webbrowser import open_new_tab
from xml.sax.saxutils import escape  # '<' -> '&lt;'

# 3rd party
from dotenv import load_dotenv
from isogeo_pysdk import IsogeoUtils, Metadata, Share
from openpyxl import load_workbook
from PyQt5.QtCore import QUrl
from PyQt5.QtWidgets import QFileDialog

# Depending on operating system
if opersys == "win32":
    """Only on MS Windows."""
    from os import startfile  # to open a folder/file
else:
    pass

# ##############################################################################
# ############ Globals ############
# #################################

if Path(".env").exists():
    load_dotenv(".env")
logger = logging.getLogger("isogeo2office")  # LOG

# ############################################################################
# ########## Classes ###############
# ##################################


class isogeo2office_utils(IsogeoUtils):
    """isogeo2office utils methods class."""

    def __init__(self):
        """Instanciating method."""
        super(isogeo2office_utils, self).__init__()

    # MISCELLANOUS -----------------------------------------------------------
    def clean_credentials_files(self, folder_to_clean: str = "_auth"):
        """Remove files from the specified folder which are older than a month.
        See: https://github.com/isogeo/isogeo-2-office/issues/51

        :param str folder_to_clean: directory to clean
        """
        # timestamp a month ago
        last_month = datetime.now() - timedelta(days=31)
        # check folder
        folder_path = Path(folder_to_clean)
        if folder_path.is_dir():
            for f in folder_path.glob("*.json"):
                if f.name == "client_secrets.json":
                    logger.debug("Credentials file in use. Do not delete it.")
                    continue
                elif f.stat().st_mtime > last_month.timestamp():
                    logger.debug(
                        "Credentials file modified during last 31 days: " + f.name
                    )
                    continue
                else:
                    logger.info(
                        "Credentials file older than 31 days marked for deletion: "
                        + f.name
                    )
                    try:
                        f.unlink()
                    except Exception as e:
                        logger.error(e)
        else:
            raise TypeError

    def proxy_settings(self):
        """Retrieves network proxy settings from OS or an environment file."""
        if getproxies():
            proxy_settings = getproxies()
            logger.debug("Proxies settings found in the OS.")
        elif environ.get("HTTP_PROXY") or environ.get("HTTPS_PROXY"):
            proxy_settings = {
                "http": environ.get("HTTP_PROXY"),
                "https": environ.get("HTTPS_PROXY"),
            }
            logger.debug(
                "Proxies settings found in environment vars (loaded from .env file)."
            )
        else:
            logger.debug("No proxy settings found.")
            proxy_settings = None

        return proxy_settings

    # -- OPENERS -----------------------------------------------------------------------
    def open_urls(self, li_url):
        """Open URLs in new tabs in the default brower.

        It waits a few seconds between the first and the next URLs
        to handle case when the webbrowser is not yet opened.

        :param list li_url: list of URLs to open in the default browser
        """
        if isinstance(li_url, QUrl):
            li_url = [li_url.toString()]
        x = 1
        for url in li_url:
            if x > 1:
                sleep(3)
            else:
                pass
            open_new_tab(url)
            x += 1

    def open_dir_file(self, target: str):
        """Open a file or a directory in the explorer of the operating system.

        :param str target: path of the folder or file to open
        """
        target_dir = Path(target)
        # check if the file or the directory exists
        if not target_dir.exists():
            raise IOError("No such file/folder: {0}".format(target))

        # check the read permission
        if not access(target, R_OK):
            raise IOError("Cannot access file: {0}".format(target))

        # open the directory or the file according to the os
        if opersys == "win32":  # Windows
            proc = startfile(target_dir.resolve())

        elif opersys.startswith("linux"):  # Linux:
            proc = subprocess.Popen(
                ["xdg-open", target_dir.resolve()],
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
            )

        elif opersys == "darwin":  # Mac:
            proc = subprocess.Popen(
                ["open", "--", target_dir.resolve()],
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
            )

        else:
            raise NotImplementedError(
                "Your `%s` isn't a supported operating system`." % opersys
            )

        # end of function
        return proc

    # UI
    def open_FileNameDialog(
        self, parent=None, file_type: str = "credentials", from_dir: str = "downloads"
    ):
        """Manage file dialog to allow user pick a file.

        :param QApplication parent: Qt parent application.
        :param str file_type: credentials | thumbnails | folder
        :param str from_dir: path to the start directory. Default value: "downloads"
        """
        # check the folder to open from
        if isinstance(from_dir, str) and from_dir.lower() == "downloads":
            # get user download directory
            start_dir = Path.home() / "Downloads"
        else:
            start_dir = Path(from_dir)
        # check if folder exists
        if not start_dir.exists():
            start_dir = Path.home()

        # set options
        options = QFileDialog.Options()
        # options |= QFileDialog.DontUseNativeDialog
        options |= QFileDialog.ReadOnly

        # adapt file filters according to file_type option
        if file_type == "credentials":
            file_filters = (
                "Standard credentials file (client_secrets.json);;JSON Files (*.json)"
            )
            dlg_title = parent.tr("Open credentials file")
            return QFileDialog.getOpenFileName(
                parent=None,
                caption=dlg_title,
                directory=str(start_dir.resolve()),
                filter=file_filters,
                options=options,
            )
        elif file_type == "thumbnails":
            file_filters = (
                "Standard credentials file (client_secrets.json);;JSON Files (*.json)"
            )
            dlg_title = parent.tr("Select thumbnails file")
            return QFileDialog.getOpenFileName(
                parent=None,
                caption=dlg_title,
                directory=str(start_dir.resolve()),
                filter=file_filters,
                options=options,
            )
        elif file_type == "folder":
            options |= QFileDialog.ShowDirsOnly
            dlg_title = parent.tr("Select folder")
            return QFileDialog.getExistingDirectory(
                parent=None,
                caption=dlg_title,
                directory=str(start_dir.resolve()),
                options=options,
            )
        else:
            file_filters = "All Files (*)"
            dlg_title = parent.tr("Pick a file")
            return QFileDialog.getOpenFileName(
                parent=None,
                caption=dlg_title,
                directory=str(start_dir.resolve()),
                filter=file_filters,
                options=options,
            )

    def timestamps_picker(self, timestamp_opt: str = "no"):
        """Return timestamp value depending on toggled radio button.

        :param str timestamp_opt: no | day | datetime
        """
        dstamp = datetime.now()
        timestamps = {
            "no": " ",
            "day": "_{}-{}-{}".format(dstamp.year, dstamp.month, dstamp.day),
            "datetime": "_{}-{}-{}-{}{}{}".format(
                dstamp.year,
                dstamp.month,
                dstamp.day,
                dstamp.hour,
                dstamp.minute,
                dstamp.second,
            ),
        }
        logger.debug(timestamp_opt)
        logger.debug(
            "Timestamp option picked: {}".format(timestamps.get(timestamp_opt))
        )
        return timestamps.get(timestamp_opt).strip()

    # ISOGEO -----------------------------------------------------------------
    @classmethod
    def get_matching_share(
        self, metadata: Metadata, shares: list, mode: str = "simple"
    ) -> Share:
        """Get the first Share which contains the metadata, basing match on workgroup UUID.

        :param Metadata metadata: metadata object to use for the matching
        :param list shares: list of shares to use for the matching
        :param str mode: simple mode is based only on the UID
        """
        if mode != "simple":
            raise NotImplementedError

        matching_share = [
            share
            for share in shares
            if share.get("_creator").get("_id") == metadata.groupId
        ]
        if len(matching_share):
            matching_share = Share(**matching_share[0])
        else:
            logger.warning(
                "No matching share found for {} ({}). The OpenCatalog URL will not be build.".format(
                    metadata.title_or_name(), metadata._id
                )
            )
            matching_share = None

        return matching_share

    def thumbnails_mngr(
        self, in_xlsx_table: str = "thumbnails/thumbnails.xlsx"
    ) -> dict:
        """Manage the thumbnails table (see: #10): check, read and return a dict.

        :param str in_xlsx_table: path to the input thumbnails table
        """
        thumbnails_dict = {}
        # check filepath
        if not path.exists(in_xlsx_table):
            logger.info("Thumbnails - Table not present. An empty one will be created.")
            return {None: (None, None)}

        # check if file is not already opened
        try:
            open(path.realpath(in_xlsx_table), "r+")
        except IOError:
            logger.error("Thumbnails table file is already opened.")
            raise IOError("Thumbnails - Table is already opened.")

        # load XLSX and check structure
        # with load_workbook(path.realpath(in_xlsx_table), read_only=True) as wb:
        wb = load_workbook(path.realpath(in_xlsx_table), read_only=True)
        if "i2o_thumbnails" not in wb.sheetnames:
            logger.error(
                "Thumbnails workbook ({}) doesn't have the good worksheet name".format(
                    in_xlsx_table
                )
            )
            wb.close()
            raise KeyError("Thumbnails - Bad worksheet name")

        # load worksheet and check headers
        ws = wb["i2o_thumbnails"]
        if not all(
            (
                ws._get_cell(1, 1).value == "isogeo_uuid",
                ws._get_cell(1, 2).value == "isogeo_title_slugged",
                ws._get_cell(1, 3).value == "img_abs_path",
            )
        ):
            logger.error(
                "Thumbnails workbook ({}) doesn't have the good headers".format(
                    in_xlsx_table
                )
            )
            raise KeyError("Thumbnails - Bad worksheet headers")

        # parse worksheet and populate final dict
        for row in ws.iter_rows(min_row=2):
            if len(row) == 3 and row[0].value and row[2].value:
                thumbnails_dict[row[0].value] = row[2].value
            else:
                logger.debug("Thumbnails reader: empty cell spotted. Quit reading.")
                break
        return thumbnails_dict

    # -- ENCODING AND NAMING STUFFS -------------------------------------------
    def clean_filename(self, filename: str, substitute: str = "", mode: str = "soft"):
        """Remove invalid characters from filename.
        \\ TO DO: isnt' duplicated with next method on special chars?

        :param str filename: filename string to clean
        :param str substitute: character to use for subtistution of special chars
        :param str modeaccents: mode to apply. Available options:

          * soft [default]: remove chars which are not accepted in filenames
          * strict: remove additional chars (punctuation...)
        """
        if mode == "soft":
            return re.sub(r'[\\/*?:"<>|]', substitute, filename)
        elif mode == "strict":
            return re.sub(r"[^\w\-_\. ]", substitute, filename)
        else:
            raise ValueError("'mode' option must be one of: soft | strict")

    def clean_special_chars(
        self, input_str: str, substitute: str = "", accents: bool = 1
    ):
        """Clean string from special characters.

        Source: https://stackoverflow.com/a/38799620/2556577

        :param str input_str: string to clean
        :param str substitute: character to use for subtistution of special chars
        :param bool accents: option to keep or not the accents
        """
        if accents:
            return re.sub(r"\W+", substitute, input_str)
        else:
            return re.sub(r"[^A-Za-z0-9]+", substitute, input_str)

    def clean_xml(self, invalid_xml: str, mode: str = "soft", substitute: str = "_"):
        """Clean string of XML invalid characters.

        source: https://stackoverflow.com/a/13322581/2556577

        :param str invalid_xml: xml string to clean
        :param str substitute: character to use for subtistution of special chars
        :param str modeaccents: mode to apply. Available options:

          * soft [default]: remove chars which are not accepted in XML
          * strict: remove additional chars
        """
        if invalid_xml is None:
            return ""
        # assumptions:
        #   doc = *( start_tag / end_tag / text )
        #   start_tag = '<' name *attr [ '/' ] '>'
        #   end_tag = '<' '/' name '>'
        ws = r"[ \t\r\n]*"  # allow ws between any token
        # note: expand if necessary but the stricter the better
        name = "[a-zA-Z]+"
        # note: fragile against missing '"'; no "'"
        attr = '{name} {ws} = {ws} "[^"]*"'
        start_tag = "< {ws} {name} {ws} (?:{attr} {ws})* /? {ws} >"
        end_tag = "{ws}".join(["<", "/", "{name}", ">"])
        tag = "{start_tag} | {end_tag}"

        assert "{{" not in tag
        while "{" in tag:  # unwrap definitions
            tag = tag.format(**vars())

        tag_regex = re.compile("(%s)" % tag, flags=re.VERBOSE)

        # escape &, <, > in the text
        iters = [iter(tag_regex.split(invalid_xml))] * 2
        pairs = zip_longest(*iters, fillvalue="")  # iterate 2 items at a time

        # get the clean version
        clean_version = "".join(escape(text) + tag for text, tag in pairs)
        if mode == "strict":
            clean_version = re.sub(r"<.*?>", substitute, clean_version)
        else:
            pass
        return clean_version


# ############################################################################
# ##### Stand alone program ########
# ##################################
if __name__ == "__main__":
    """Standalone execution and tests"""
    utils = isogeo2office_utils()

    utils.clean_credentials_files(r"_auth")
