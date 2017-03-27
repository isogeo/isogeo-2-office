# -*- coding: UTF-8 -*-
#!/usr/bin/env python
from __future__ import (absolute_import, print_function, unicode_literals)
# ------------------------------------------------------------------------------
# Name:         OpenCatalog to Excel
# Purpose:      Get metadatas from an Isogeo OpenCatalog and store it into
#               an Excel workbook (.xls).
#
# Author:       Julien Moura (@geojulien) & Valentin Blanlot (@bablot)
#
# Python:       2.7.x
# Created:      14/08/2014
# Updated:      12/12/2015
# ------------------------------------------------------------------------------

# ##############################################################################
# ########## Libraries #############
# ##################################

# Standard library

# 3rd party library
from openpyxl import load_workbook

# ##############################################################################s
# ########## Classes ###############
# ##################################


class XslxReader(object):
    """ Abstraction class to manipulate XLSX input files and avoid
    repeating actions
    """
    def __init__(self, xlsx_path, sheet_idx=0):
        """ XLSX

        Keyword arguments:

        """
        super(XslxReader, self).__init__()

        # somes checks
        self.check_input_xlsx(xlsx_path)

        # loading input file
        xlsx_in = load_workbook(filename=xlsx_path,
                                read_only=True,
                                guess_types=True,
                                data_only=True)

        # return the requested worksheet
        self.ws = xlsx_in.worksheets[sheet_idx]

    def check_input_xlsx(self, xlsx_path):
        """ Check some basic parameters
        """

        # end of method
        return

    def get_headers_names(self, start_row=1):
        """ Get columns names
        """
        # headers row
        if start_row != 1:
            header = start_row
        else:
            header = self.ws.min_row

        return [self.ws.cell(row=header, column=col).value for col in range(1, self.ws.max_column)]




###############################################################################
###### Stand alone program ########
###################################

if __name__ == '__main__':
    """ standalone execution """
    xlsheet = XslxReader(r"..\input\wb_test.xlsx")

    print(xlsheet.get_headers_names())
